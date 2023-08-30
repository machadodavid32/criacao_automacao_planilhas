import openpyxl

'''
Quando falo de 
workbook, falo de planilha

Quando falo de
sheet, falo de página'''

# Criação de uma planilha
workbook = openpyxl.Workbook()

# mostrar sheets(paginas) existentes
print(workbook.sheetnames)
# Resposta: ['Sheet']


# Criando sheets
workbook.create_sheet('ruas') # nome da sheet(pagina)
workbook.create_sheet('cidades')
workbook.create_sheet('estados')
# salvar modificações e finalizando processo de criaçãp
workbook.save('endereços.xlsl') # nome da planilha


# alterar o nome de um sheet
workbook['ruas'].title = 'Ruas da Cidade'
workbook.save('endereços.xlsl')


# Deletar um sheet da minha planilha
del workbook('Sheet')
print(workbook.sheetnames)
workbook.save('endereços.xlsl')


'''alteração
'''