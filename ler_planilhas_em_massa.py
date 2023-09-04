import openpyxl

workbook = openpyxl.load_workbook('hockey-players.xlsx') # caso este arquivo tivesse em algum outro lugar do computador, passar caminho completo
sheet_player_data = workbook['PlayerData'] # planilha PlayerData
print(sheet_player_data['D3'].value) # vai imprimir a informação na celula escolhida
sheet_player_data['D3'].value = 'Amanda'  # vai substituir a informação na celula e colocar esta nova.
workbook.save('hocker_players_novo.xlsx')

for linha in sheet_player_data.iter_rows(min_row=2, max_row=5, min_col=1, max_col=11): 
# Acima quer dizer que vai da linha 2 até 5 e da coluna 1 até 11 do excel.
    print(linha[0].value, linha[1].value, linha[2].value, linha[3].value, linha[4].value, linha[5].value,
          linha[6].value, linha[7].value, linha[8].value, linha[9].value, linha[10].value)
    # Acima serve somente para printar o resultado no terminal.

for coluna in sheet_player_data.iter_cols(min_col=1, max_col=3, min_row=2, max_row=5):
    # Acima estamos iterando sobre as colunas, neste caso, coluna 1 até 3 e linha 2 até 5
    for celula in coluna:
        print(celula.value)
            
